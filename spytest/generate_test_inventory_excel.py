#!/usr/bin/env python3
"""
Generate Test Inventory Excel from batch_full_run.sh

This script:
1. Parses batch_full_run.sh to extract all test scripts
2. Analyzes each test script for metadata
3. Parses execution logs (if available)
4. Generates comprehensive Excel report

Output columns:
- Script Name
- Script Relative Location (from tests/)
- Feature
- CLI Type
- Total Testcases
- Passed
- Failed
- Defect ID (if any)

Usage:
    python3 generate_test_inventory_excel.py
    python3 generate_test_inventory_excel.py --log-root ./logs/20260226
    python3 generate_test_inventory_excel.py --output test_inventory.xlsx
"""

import argparse
import re
import os
import sys
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import ast

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Install with: pip3 install openpyxl")
    sys.exit(1)


class TestScriptAnalyzer:
    """Analyzes test scripts to extract metadata"""

    def __init__(self, script_path: Path):
        self.script_path = script_path
        self.content = ""
        self.lines = []

        if script_path.exists():
            try:
                with open(script_path, 'r', encoding='utf-8') as f:
                    self.content = f.read()
                    self.lines = self.content.split('\n')
            except Exception as e:
                print(f"WARNING: Could not read {script_path}: {e}")

    def get_cli_type(self) -> str:
        """Extract CLI type from script"""
        # Check docstring
        if "klish" in self.content.lower():
            if "click" in self.content.lower():
                return "klish,click"
            return "klish"
        elif "click" in self.content.lower():
            return "click"

        # Check for st.get_ui_type() calls
        if 'cli_type="klish"' in self.content or "cli_type='klish'" in self.content:
            return "klish"
        if 'cli_type="click"' in self.content or "cli_type='click'" in self.content:
            return "click"

        # Check for REST/gNMI
        if "rest" in self.content.lower():
            return "REST"
        if "gnmi" in self.content.lower():
            return "gNMI"

        return "mixed/auto"

    def get_feature_from_docstring(self) -> Optional[str]:
        """Extract feature from docstring"""
        # Look for patterns like "Feature: BGP" or "Test Case: ..."
        feature_patterns = [
            r'Feature[:\s]+(\w+)',
            r'Test Case[:\s]+(.+?)(?:\n|$)',
            r'Test ID[:\s]+(\S+)',
        ]

        for pattern in feature_patterns:
            match = re.search(pattern, self.content, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).strip()

        return None

    def count_test_functions(self) -> int:
        """Count test functions in script"""
        try:
            tree = ast.parse(self.content)
            test_funcs = [
                node for node in ast.walk(tree)
                if isinstance(node, ast.FunctionDef) and node.name.startswith('test_')
            ]
            return len(test_funcs)
        except:
            # Fallback to regex
            pattern = r'^\s*def\s+(test_\w+)\s*\('
            matches = re.findall(pattern, self.content, re.MULTILINE)
            return len(matches)

    def get_test_function_names(self) -> List[str]:
        """Get list of test function names"""
        try:
            tree = ast.parse(self.content)
            test_funcs = [
                node.name for node in ast.walk(tree)
                if isinstance(node, ast.FunctionDef) and node.name.startswith('test_')
            ]
            return test_funcs
        except:
            pattern = r'^\s*def\s+(test_\w+)\s*\('
            matches = re.findall(pattern, self.content, re.MULTILINE)
            return matches


class BatchFileParser:
    """Parses batch_full_run.sh to extract test information"""

    def __init__(self, batch_file: Path):
        self.batch_file = batch_file
        self.batches = {}
        self.parse()

    def parse(self):
        """Parse batch_full_run.sh"""
        with open(self.batch_file, 'r') as f:
            content = f.read()

        # Extract batch definitions
        batch_pattern = r'run_(?:bgp_)?batch\s+"([^"]+)"\s+"([^"]+)"\s+\\?\s*((?:[^\n]*\\\s*\n)*[^\n]*)'

        matches = re.finditer(batch_pattern, content, re.MULTILINE)

        for match in matches:
            feature_name = match.group(1)
            testbed = match.group(2)
            tests_block = match.group(3)

            # Extract individual test files
            test_files = []
            for line in tests_block.split('\n'):
                line = line.strip().rstrip('\\').strip()
                if line and not line.startswith('#') and '.py' in line:
                    # Clean up the path
                    test_file = line.strip()
                    if test_file:
                        test_files.append(test_file)

            if feature_name not in self.batches:
                self.batches[feature_name] = {
                    'testbed': testbed,
                    'tests': []
                }

            self.batches[feature_name]['tests'].extend(test_files)


class LogParser:
    """Parses SPyTest log files to extract test results"""

    def __init__(self, log_root: Optional[Path]):
        self.log_root = log_root
        self.results = {}

    def find_summary_file(self, feature: str) -> Optional[Path]:
        """Find summary.txt for a feature"""
        if not self.log_root or not self.log_root.exists():
            return None

        # Look for feature directories
        feature_dirs = list(self.log_root.glob(f"{feature}/*"))
        if not feature_dirs:
            return None

        # Get most recent
        feature_dir = sorted(feature_dirs, reverse=True)[0]
        summary_file = feature_dir / "summary.txt"

        return summary_file if summary_file.exists() else None

    def parse_summary(self, summary_file: Path) -> Dict[str, Dict]:
        """Parse summary.txt to extract test results"""
        results = {}

        if not summary_file.exists():
            return results

        try:
            with open(summary_file, 'r') as f:
                content = f.read()

            # Look for result lines like:
            # test_static_route_basic.py::test_static_route ... PASS
            # test_static_route_basic.py::test_static_route ... FAIL
            pattern = r'(\S+\.py)::(\S+)\s+\.\.\.\s+(PASS|FAIL|SKIP|UNSUPPORTED|ENVFAIL)'

            for match in re.finditer(pattern, content):
                script = match.group(1)
                test_func = match.group(2)
                result = match.group(3)

                if script not in results:
                    results[script] = {
                        'total': 0,
                        'passed': 0,
                        'failed': 0,
                        'skipped': 0,
                        'other': 0,
                        'tests': {}
                    }

                results[script]['total'] += 1
                results[script]['tests'][test_func] = result

                if result == 'PASS':
                    results[script]['passed'] += 1
                elif result == 'FAIL':
                    results[script]['failed'] += 1
                elif result in ('SKIP', 'UNSUPPORTED'):
                    results[script]['skipped'] += 1
                else:
                    results[script]['other'] += 1

        except Exception as e:
            print(f"WARNING: Could not parse {summary_file}: {e}")

        return results


class ExcelGenerator:
    """Generates Excel file with test inventory"""

    def __init__(self, output_file: Path):
        self.output_file = output_file
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Test Inventory"

        # Define styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Status colors
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def write_headers(self):
        """Write column headers"""
        headers = [
            "Batch",
            "Script Name",
            "Script Relative Location",
            "Feature",
            "CLI Type",
            "Total Testcases",
            "Passed",
            "Failed",
            "Skipped",
            "Pass Rate %",
            "Defect ID",
            "Notes"
        ]

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Freeze header row
        self.ws.freeze_panes = "A2"

    def add_test_row(self, row_num: int, data: Dict):
        """Add a test row to the worksheet"""
        columns = [
            data.get('batch', ''),
            data.get('script_name', ''),
            data.get('script_location', ''),
            data.get('feature', ''),
            data.get('cli_type', ''),
            data.get('total', 0),
            data.get('passed', 0),
            data.get('failed', 0),
            data.get('skipped', 0),
            data.get('pass_rate', ''),
            data.get('defect_id', ''),
            data.get('notes', '')
        ]

        for col, value in enumerate(columns, 1):
            cell = self.ws.cell(row=row_num, column=col)
            cell.value = value
            cell.border = self.border

            if col in (6, 7, 8, 9):  # Numeric columns
                cell.alignment = Alignment(horizontal='center')

            if col == 10:  # Pass rate
                cell.alignment = Alignment(horizontal='center')

        # Color code based on results
        if data.get('failed', 0) > 0:
            self.ws.cell(row=row_num, column=8).fill = self.fail_fill
        elif data.get('passed', 0) > 0 and data.get('failed', 0) == 0:
            self.ws.cell(row=row_num, column=7).fill = self.pass_fill

    def adjust_column_widths(self):
        """Auto-adjust column widths"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def add_summary_sheet(self, batch_summary: Dict):
        """Add summary worksheet"""
        ws = self.wb.create_sheet("Summary")

        # Title
        ws['A1'] = "Test Execution Summary"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "Batch"
        ws[f'B{row}'] = "Total Scripts"
        ws[f'C{row}'] = "Total Testcases"
        ws[f'D{row}'] = "Passed"
        ws[f'E{row}'] = "Failed"
        ws[f'F{row}'] = "Pass Rate %"

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        row += 1
        for batch, stats in sorted(batch_summary.items()):
            ws[f'A{row}'] = batch
            ws[f'B{row}'] = stats['scripts']
            ws[f'C{row}'] = stats['total']
            ws[f'D{row}'] = stats['passed']
            ws[f'E{row}'] = stats['failed']

            pass_rate = 0
            if stats['total'] > 0:
                pass_rate = (stats['passed'] / stats['total']) * 100
            ws[f'F{row}'] = f"{pass_rate:.1f}%"

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border

            row += 1

        # Adjust widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    def save(self):
        """Save workbook"""
        self.wb.save(self.output_file)
        print(f"Excel file saved: {self.output_file}")


def main():
    parser = argparse.ArgumentParser(description="Generate Test Inventory Excel from batch_full_run.sh")
    parser.add_argument('--batch-file', default='./batch_full_run.sh',
                       help='Path to batch_full_run.sh (default: ./batch_full_run.sh)')
    parser.add_argument('--log-root', type=str,
                       help='Root directory of test logs (e.g., ./logs/20260226)')
    parser.add_argument('--output', '-o', default='test_inventory.xlsx',
                       help='Output Excel file (default: test_inventory.xlsx)')
    parser.add_argument('--tests-dir', default='./tests',
                       help='Tests directory (default: ./tests)')

    args = parser.parse_args()

    batch_file = Path(args.batch_file)
    if not batch_file.exists():
        print(f"ERROR: Batch file not found: {batch_file}")
        sys.exit(1)

    tests_dir = Path(args.tests_dir)
    if not tests_dir.exists():
        print(f"ERROR: Tests directory not found: {tests_dir}")
        sys.exit(1)

    log_root = Path(args.log_root) if args.log_root else None

    print("=" * 60)
    print(" Test Inventory Excel Generator")
    print("=" * 60)
    print(f"Batch file : {batch_file}")
    print(f"Tests dir  : {tests_dir}")
    print(f"Log root   : {log_root if log_root else 'None (template only)'}")
    print(f"Output file: {args.output}")
    print("=" * 60)

    # Parse batch file
    print("\n[1/4] Parsing batch_full_run.sh...")
    batch_parser = BatchFileParser(batch_file)
    print(f"      Found {len(batch_parser.batches)} batches")
    for batch_name, batch_info in batch_parser.batches.items():
        print(f"      - {batch_name}: {len(batch_info['tests'])} tests")

    # Parse logs if available
    print("\n[2/4] Parsing test logs...")
    log_parser = LogParser(log_root)
    all_log_results = {}

    if log_root:
        for feature in batch_parser.batches.keys():
            summary_file = log_parser.find_summary_file(feature)
            if summary_file:
                print(f"      Found summary for {feature}: {summary_file}")
                results = log_parser.parse_summary(summary_file)
                all_log_results.update(results)
        print(f"      Parsed results for {len(all_log_results)} test scripts")
    else:
        print("      No log root specified - generating template only")

    # Analyze test scripts
    print("\n[3/4] Analyzing test scripts...")
    excel = ExcelGenerator(Path(args.output))
    excel.write_headers()

    row_num = 2
    batch_summary = {}

    for batch_name, batch_info in sorted(batch_parser.batches.items()):
        if batch_name not in batch_summary:
            batch_summary[batch_name] = {
                'scripts': 0,
                'total': 0,
                'passed': 0,
                'failed': 0,
                'skipped': 0
            }

        for test_file in batch_info['tests']:
            # Clean up path
            test_file = test_file.strip()
            script_name = Path(test_file).name

            # Full path to script
            script_path = tests_dir / test_file

            print(f"      Analyzing: {test_file}")

            # Analyze script
            analyzer = TestScriptAnalyzer(script_path)
            cli_type = analyzer.get_cli_type()
            feature = analyzer.get_feature_from_docstring()

            # If no feature from docstring, derive from path
            if not feature:
                parts = test_file.split('/')
                if len(parts) > 1:
                    feature = parts[0].upper()
                else:
                    feature = "Unknown"

            total_tests = analyzer.count_test_functions()

            # Get results from logs
            passed = 0
            failed = 0
            skipped = 0

            if script_name in all_log_results:
                log_data = all_log_results[script_name]
                total_tests = log_data['total']  # Use actual from logs
                passed = log_data['passed']
                failed = log_data['failed']
                skipped = log_data['skipped']

            # Calculate pass rate
            pass_rate = ""
            if total_tests > 0 and (passed + failed) > 0:
                pass_rate = f"{(passed / (passed + failed)) * 100:.1f}%"

            # Update batch summary
            batch_summary[batch_name]['scripts'] += 1
            batch_summary[batch_name]['total'] += total_tests
            batch_summary[batch_name]['passed'] += passed
            batch_summary[batch_name]['failed'] += failed
            batch_summary[batch_name]['skipped'] += skipped

            # Add row
            row_data = {
                'batch': batch_name,
                'script_name': script_name,
                'script_location': test_file,
                'feature': feature,
                'cli_type': cli_type,
                'total': total_tests,
                'passed': passed,
                'failed': failed,
                'skipped': skipped,
                'pass_rate': pass_rate,
                'defect_id': '',
                'notes': ''
            }

            excel.add_test_row(row_num, row_data)
            row_num += 1

    # Add summary sheet
    print("\n[4/4] Generating Excel file...")
    excel.add_summary_sheet(batch_summary)
    excel.adjust_column_widths()
    excel.save()

    print("\n" + "=" * 60)
    print(" Summary")
    print("=" * 60)
    print(f"Total Batches      : {len(batch_parser.batches)}")
    print(f"Total Test Scripts : {row_num - 2}")
    print(f"Output File        : {args.output}")
    print("=" * 60)

    # Display batch summary
    print("\nBatch Summary:")
    print("-" * 80)
    print(f"{'Batch':<30} {'Scripts':>8} {'Tests':>8} {'Passed':>8} {'Failed':>8}")
    print("-" * 80)
    for batch, stats in sorted(batch_summary.items()):
        print(f"{batch:<30} {stats['scripts']:>8} {stats['total']:>8} "
              f"{stats['passed']:>8} {stats['failed']:>8}")
    print("-" * 80)

    print("\nDone!")


if __name__ == '__main__':
    main()
